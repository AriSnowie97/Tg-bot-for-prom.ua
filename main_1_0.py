import os
import logging
import tempfile
import uuid
import openpyxl
from openpyxl import Workbook
from telegram import Update, InputFile
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters

# 🔧 Настройка логгирования
logging.basicConfig(level=logging.INFO)

# 🟢 Стартовая команда
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет! Отправь строку с ссылками (через запятую) или Excel-файл (.xlsx).\n"
        "Я верну тебе первые 10 ссылок по каждой строке."
    )

# 🔤 Обработка текста
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message.text
    image_urls = [url.strip() for url in message.split(",") if url.strip().startswith("http")]
    first_10 = image_urls[:10]

    if not first_10:
        await update.message.reply_text("❗️Не найдено валидных ссылок.")
        return

    response = "🔟 Первые 10 ссылок:\n\n" + "\n".join(first_10)
    await update.message.reply_text(response)

# 📁 Обработка Excel-файлов
async def handle_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document

    if not document.file_name.endswith(".xlsx"):
        await update.message.reply_text("❌ Пожалуйста, отправь Excel-файл с расширением .xlsx")
        return

    try:
        temp_filename = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4()}.xlsx")
        file = await document.get_file()
        await file.download_to_drive(temp_filename)

        wb = openpyxl.load_workbook(temp_filename)
        sheet = wb.active

        messages = []
        output_data = []

        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_col=1), start=1):
            cell = row[0]
            if cell.value:
                parts = str(cell.value).split(",")
                links = [url.strip() for url in parts if url.strip().startswith("http")]
                top_links = links[:10]

                if top_links:
                    text = f"🧾 Строка {idx}: {len(top_links)} ссылок\n" + "\n".join(top_links)
                    messages.append(text)
                    output_data.append([idx, len(top_links), "\n".join(top_links)])
                else:
                    messages.append(f"🧾 Строка {idx}: нет валидных ссылок")
                    output_data.append([idx, 0, ""])

        # Сообщения
        if messages:
            for chunk in split_messages(messages):
                await update.message.reply_text(chunk)
        else:
            await update.message.reply_text("⚠️ В файле не найдено ни одной строки с валидными ссылками.")

        # Создание выходного Excel
        if output_data:
            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.title = "Результаты"
            out_ws.append(["Строка", "Кол-во ссылок", "Ссылки (до 10)"])
            for row in output_data:
                out_ws.append(row)

            result_path = os.path.join(tempfile.gettempdir(), f"result_{uuid.uuid4()}.xlsx")
            out_wb.save(result_path)

            with open(result_path, "rb") as f:
                await update.message.reply_document(
                    document=InputFile(f, filename="processed_links.xlsx"),
                    caption="📄 Обработанный файл с ссылками"
                )
            os.remove(result_path)

    except Exception as e:
        logging.error(f"Ошибка при обработке Excel: {e}")
        await update.message.reply_text("❌ Произошла ошибка при чтении файла.")
    finally:
        if os.path.exists(temp_filename):
            os.remove(temp_filename)

# 📜 Разбиение длинных сообщений
def split_messages(messages, limit=4000):
    chunks = []
    current = ""

    for msg in messages:
        if len(current) + len(msg) + 2 > limit:
            chunks.append(current)
            current = msg
        else:
            if current:
                current += "\n\n" + msg
            else:
                current = msg

    if current:
        chunks.append(current)

    return chunks

# ▶️ Запуск бота
def main():
    TOKEN = "7575947696:AAEAuezyu_AEnpG2IMCPXFk6i_lW9zAos1w"

    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(
        filters.Document.MimeType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        handle_excel
))


    print("🤖 Бот запущен!")
    app.run_polling()

if __name__ == "__main__":
    main()
